import os

import docx
import fitz  # PyMuPDF
from openpyxl import load_workbook


def extract_text(file_path: str) -> str:
    """Extract text from a PDF, DOCX, TXT, Excel, CSV, or Markdown file."""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".pdf":
        return _extract_pdf(file_path)
    elif ext == ".docx":
        return _extract_docx(file_path)
    elif ext == ".txt":
        return _extract_txt(file_path)
    elif ext == ".xlsx":
        return _extract_xlsx(file_path)
    elif ext == ".xls":
        return _extract_xls(file_path)
    elif ext == ".csv":
        return _extract_csv(file_path)
    elif ext == ".md":
        return _extract_markdown(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def _extract_pdf(path: str) -> str:
    doc = fitz.open(path)
    text = ""
    for page in doc:
        text += page.get_text()
    doc.close()
    return text


def _extract_docx(path: str) -> str:
    doc = docx.Document(path)
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def _extract_xlsx(path: str) -> str:
    """Extract text from Excel .xlsx file."""
    workbook = load_workbook(path)
    text = []
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        text.append(f"Sheet: {sheet_name}\n")
        
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                text.append(row_text)
    
    return "\n".join(text)


def _extract_xls(path: str) -> str:
    """Extract text from Excel .xls file."""
    try:
        import xlrd
        workbook = xlrd.open_workbook(path)
        text = []
        
        for sheet in workbook.sheets():
            text.append(f"Sheet: {sheet.name}\n")
            
            for row in range(sheet.nrows):
                row_values = sheet.row_values(row)
                row_text = " | ".join(str(cell) if cell else "" for cell in row_values)
                if row_text.strip():
                    text.append(row_text)
        
        return "\n".join(text)
    except ImportError:
        return f"Error: xlrd library not installed. Cannot extract .xls files. Please install it with: pip install xlrd"


def _extract_csv(path: str) -> str:
    """Extract text from CSV file."""
    import csv
    text = []
    
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        for row in reader:
            row_text = " | ".join(cell.strip() for cell in row)
            if row_text.strip():
                text.append(row_text)
    
    return "\n".join(text)


def _extract_markdown(path: str) -> str:
    """Extract text from Markdown file."""
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()
